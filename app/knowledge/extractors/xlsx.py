from ..models import ExtractedDocument, ExtractedUnit, SourceMetadata
from .base import ExtractionError, ExtractionLimits, enforce_limits


def extract(path: str, source: SourceMetadata, limits: ExtractionLimits) -> ExtractedDocument:
    try:
        from openpyxl import load_workbook
        workbook = load_workbook(path, read_only=True, data_only=False)
    except Exception as exc:
        raise ExtractionError(f"could not open XLSX: {exc}") from exc
    units = []
    try:
        for sheet in workbook.worksheets:
            rows, last_row, last_col = [], 0, 0
            for row_number, row in enumerate(sheet.iter_rows(), 1):
                values = ["" if cell.value is None else str(cell.value) for cell in row]
                if any(values):
                    last_row = row_number
                    last_col = max(last_col, max((index + 1 for index, value in enumerate(values) if value), default=0))
                    rows.append("\t".join(values).rstrip())
            end = f"{sheet.cell(last_row or 1, last_col or 1).coordinate}"
            units.append(ExtractedUnit("sheet", sheet.title, "\n".join(rows), end, heading=sheet.title))
    finally:
        workbook.close()
    return enforce_limits(ExtractedDocument(source.display_name, "spreadsheet", units,
                                             {"sheet_count": len(units)}), limits)
